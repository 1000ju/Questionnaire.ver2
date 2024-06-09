import random
import openpyxl


# b_gender, b_age -> 문자라서 오류가 발생했던 거임 / 하단의 코드는 시트 좌표를 찍어 값을 넣는 형태. 즉, 숫자로 관리하기 때문
def College_Age(Total, fpath, b_gender, b_age, b_grade, b_major, a_gender, a_age, a_grade, a_major):
    start_row = 4
    wb = openpyxl.load_workbook(fpath)
    b_sheet = wb["사전설문지"]
    a_sheet = wb["사후설문지"]

    # 전공*, 성별*, 나이*, 학년*,
    for i in range(Total):
        # ws.cell(row=row, column=column, value=value)
        now_row = start_row + i

        # 전공
        Major_temp = random.randrange(1, 6)
        b_sheet.cell(row=now_row, column=b_major, value=Major_temp)
        a_sheet.cell(row=now_row, column=a_major, value=Major_temp)

        # 성별
        Gender_temp = random.randrange(1, 3)
        b_sheet.cell(row=now_row, column=b_gender, value=Gender_temp)
        a_sheet.cell(row=now_row, column=a_gender, value=Gender_temp)

        # 나이
        Age_temp = random.randrange(19, 24)
        b_sheet.cell(row=now_row, column=b_age, value=Age_temp)
        a_sheet.cell(row=now_row, column=a_age, value=Age_temp)

        # 학년
        # 해당 나이에 진학이 가능한 학년 배정

        if Age_temp == 19:
            Grade_temp = 1
        elif Age_temp == 20:
            Grade_temp = random.randrange(1, 3)

        elif Age_temp == 21:
            Grade_temp = random.randrange(1, 4)

        elif Age_temp >= 22:
            Grade_temp = random.randrange(1, 5)

        b_sheet.cell(row=now_row, column=b_grade, value=Grade_temp)
        a_sheet.cell(row=now_row, column=a_grade, value=Grade_temp)

    print("사전, 사후 설문지의 나이 / 성별 / 학년/ 전공을 동시에 생성하였습니다")
    wb.save(fpath)  # 변경된 부분
    return


# 추가한 내용***
def Highschool_Age(Total, freshman, sophomore, junior, fpath, b_gender, b_age, b_grade, a_gender, a_age, a_grade):
    start_row = 3
    wb = openpyxl.load_workbook(fpath)
    b_sheet = wb["사전설문지"]
    a_sheet = wb["사후설문지"]

    # 성별*, 나이*, 학년*,
    for i in range(1, Total + 1):  # 1포함, Total+1미포함(Total까지)
        now_row = start_row + i

        # 성별 -> 상황에 따라 남/여 정확한 숫자가 주어지기도 함
        Gender_temp = random.randrange(1, 3)
        b_sheet.cell(row=now_row, column=b_gender, value=Gender_temp)
        a_sheet.cell(row=now_row, column=a_gender, value=Gender_temp)

        # 나이 / 학년
        if now_row <= freshman + start_row:
            Age_temp = 17
            Grade_temp = 1
        elif now_row > freshman and now_row <= freshman+sophomore + start_row:
            Age_temp = 18
            Grade_temp = 2
        elif now_row > sophomore and now_row <= freshman + sophomore + junior + start_row:
            Age_temp = 19
            Grade_temp = 3
        b_sheet.cell(row=now_row, column=b_age, value=Age_temp)
        a_sheet.cell(row=now_row, column=a_age, value=Age_temp)
        # 학년
        b_sheet.cell(row=now_row, column=b_grade, value=Grade_temp)
        a_sheet.cell(row=now_row, column=a_grade, value=Grade_temp)

    print("사전, 사후 설문지의 나이 / 성별 / 학년을 동시에 생성하였습니다")
    wb.save(fpath)  # 변경된 부분
    return
